from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
from .models import *

from django.db.models import Q

import json

sciences = {
    '1': {
        'Akademik yozuv': 30,
        'Ingliz tili': 30,
        'Xisob': 30,
        'Fizika': 30,
        'Dasturlash': 30,
        'Falsafa': 30
    },
    '2': {
        'Akademik yozuv': 15,
        'Ingliz tili': 30,
        'Differensial tenglamalar': 30,
        'Chiziqli algebra': 30,
        'Fizika': 30,
        'Dasturlash': 30
    },
    '3': {
        'Kiberxavfsizlik asoslari': 37,
        "Ma'lumotlar bazasi": 37,
        "Ma'lumotlar tuzilmasi va algoritmlar": 37,
        'Diskret tuzilmalar': 37,
        'Elektronika va sxemalar': 37,
        'Dasturlash': 37
    },
    '4': {
        'Kompyuterning tashkil topishi': 37,
        "Web dasturlash": 37,
        'Algoritmlarni loyihalash': 37,
        'Extimollik': 37
    },
    '5': {
        'Kompyuter tarmoqlari': 45,
        "Dasturiy ta'minotni tizimlarini loyihalash": 45,
        "Inson - kompyuter o'zaro ta'siri": 45,
        'Dasturlash usullari va paratigmalar': 45,
        'Zamonaviy iqtisodiyot': 30,
    },
    '6': {
        "Operatsion tizimlar": 45,
        "Dasturiy ta'minot arxitekturasi": 45,
        "Dasturiy ta'minot sifatini ta'minlash": 45,
        "Mobil ilovalarni ishlab chiqish": 45,
        "Zamonaviy iqtisodiyot": 30,
    },
    '7': {
        "Dasturiy ta'minot qurilmasi va evolyutsiya": 45,
        "Dasturiy ta'minot loyihalarini boshqarish": 45,
        "Sql dasturlash tili": 45,
        "Python tilida dasturlash": 45,
    },
    '8': {
        "Dasturiy vositalarni testlash": 45,
        "MATLAB da dasturlash": 45,
        "Timsollarni tanib olish": 45,
    },

}

def update(r):
    for i in sciences:
        for x, y in sciences[i].items():
            sc = Science.objects.get(semester=i, name=x)
            sc.lessons_count = y

            sc.save()

    return HttpResponse('<h1>awdawda</h1>')


def ret_marks(people):
    marks = {
        '1': {},
        '2': {},
        '3': {},
        '4': {},
        '5': {},
        '6': {},
        '7': {},
        '8': {},
    }

    for s in sciences:
        for sc in sciences[s]:
            mks = sum(x.mark for x in Mark.objects.filter(Q(science__semester=s) & Q(science__name=sc) & Q(user=people)))
            marks[s][sc] = 0 if mks == 0 else int(mks / ( sciences[s][sc] * 5 / 100 ))

    return json.dumps(marks)


def clean_string(text):
    return text.replace("'", "").lower()


def read_file(user, file, semester):

    workbook = load_workbook(filename=file)
    sheet = workbook.active
    sdata = sciences[semester]
    row, col = 2, 2

    for s, c in sdata.items():
        try:
            science, _ = Science.objects.get_or_create(
                name=s,
                semester=semester
            )

            for i in range(col, col + c):
                mark, _ = Mark.objects.get_or_create(
                    user=user, science=science, lesson_index=i - 1)
                mk = sheet.cell(row=row, column=i).value
                mark.mark = int(mk) if mk is not None else 0
                mark.save()

            row += 1

        except Exception as e:
            print(f'Error: {e}')


@login_required(login_url='login')
def home_page(request):
    if request.user.user_type == '1':
        if request.method == 'POST':
            type = request.POST.get('type')

            if type == 'add-people':
                first_name = request.POST.get('first_name')
                last_name = request.POST.get('last_name')

                user = User.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    username=f'{clean_string(last_name)}-{clean_string(first_name)}',
                    user_type='2'
                )

                user.set_password('123456789')
                user.save()
                
                return redirect('home')

            elif type == "upload-marks":
                user = User.objects.get(pk=request.POST.get('user-id'))
                file = request.FILES.get('excel-file')
                semester = request.POST.get('semester')

                read_file(user, file, semester)
                return redirect('home')

        context = {
            'peoples': User.objects.filter(user_type='2')
        }

        return render(request, 'teacher-index.html', context)
    else:
        context = {
            'semesters': [1, 2, 3, 4, 5, 6, 7, 8]
        }
        context['marks'] = ret_marks(request.user)

        return render(request, 'index.html', context)


def login_page(request):
    context = {}

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            context['error'] = 'User not found'

    return render(request, 'login.html', context)


def people_page(request, pk):
    context = {
        'semesters': [1, 2, 3, 4, 5, 6, 7, 8]
    }

    try:
        people = User.objects.get(pk=pk)

        context['marks'] = ret_marks(people)
        context['people'] = people

    except Exception as e:
        context['error'] = 'User not found'
        print(f'Error {e}')

    return render(request, 'people.html', context)


def logout_page(request):
    logout(request)
    return redirect('login')
